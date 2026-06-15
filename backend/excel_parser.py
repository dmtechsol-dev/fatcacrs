from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from backend.config import (
    EXPECTED_COLUMNS,
    EXPECTED_SHEET,
    FINANCIAL_INSTITUTION_IN_COLUMNS,
    REQUIRED_ACCOUNT_FIELDS,
)
from backend.financial_institution import normalize_financial_institution_in
from backend.models import AccountRecord, StatusMappingInfo


TRUE_VALUES = {"true", "yes", "y", "1"}
FALSE_VALUES = {"false", "no", "n", "0"}


@dataclass(frozen=True)
class WorkbookParseResult:
    records: list[AccountRecord]
    status_mapping: StatusMappingInfo
    financial_institution_in: str | None = None


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return " ".join(_text(value).lower().split())


def normalize_date(value: Any, workbook_epoch=None) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and workbook_epoch is not None:
        try:
            return from_excel(value, workbook_epoch).date().isoformat()
        except (TypeError, ValueError, OverflowError):
            return _text(value)
    raw = _text(value)
    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return raw


def map_header(value: Any) -> str | None:
    normalized = normalize_header(value)
    if normalized.startswith("account balance"):
        return "account_balance"
    return EXPECTED_COLUMNS.get(normalized)


def normalize_status_indicator(value: Any, label: str) -> tuple[bool, str]:
    raw = _text(value).lower()
    if raw in TRUE_VALUES or raw == label:
        return True, ""
    if raw in FALSE_VALUES or not raw:
        return False, ""
    return False, (
        f"Unsupported {label} status value '{_text(value)}'. "
        "Use Yes/No, Y/N, TRUE/FALSE, 1/0, or the status name."
    )


def normalize_account_status(value: Any) -> tuple[bool, str]:
    raw = _text(value).lower()
    if raw in TRUE_VALUES:
        return True, ""
    if raw in FALSE_VALUES:
        return False, ""
    shown_value = _text(value) or "(blank)"
    return False, (
        f"Invalid dormant Account Status value '{shown_value}'. "
        "Use Yes/No, Y/N, TRUE/FALSE, or 1/0."
    )


def status_mapping_info(
    mapped_headers: dict[int, str],
    header_values: tuple[Any, ...],
) -> StatusMappingInfo:
    detected = {
        field: _text(header_values[index])
        for index, field in mapped_headers.items()
        if field
        in {
            "account_status",
            "dormant_account",
            "closed_account",
            "undocumented_account",
        }
    }
    warnings: list[str] = []
    if not detected:
        warnings.append(
            "No dormant status column was detected."
        )
    else:
        for field, label in (
            ("closed_account", "ClosedAccount"),
            ("undocumented_account", "UndocumentedAccount"),
        ):
            if field not in detected:
                warnings.append(
                    f"No dedicated {label} column was detected. It defaults "
                    "to false."
                )
    return StatusMappingInfo(
        account_status=detected.get("account_status"),
        dormant_account=detected.get("dormant_account"),
        closed_account=detected.get("closed_account"),
        undocumented_account=detected.get("undocumented_account"),
        warnings=warnings,
    )


def is_summary_row(values: dict[str, Any]) -> bool:
    if _text(values.get("first_name")) or _text(values.get("surname")):
        return False
    combined = " ".join(_text(value) for value in values.values()).lower()
    return "true:" in combined or "false:" in combined


def parse_excel_with_metadata(content: bytes) -> WorkbookParseResult:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if EXPECTED_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"Workbook must contain a sheet named '{EXPECTED_SHEET}'."
        )

    sheet = workbook[EXPECTED_SHEET]
    header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapped_headers: dict[int, str] = {}
    financial_institution_in_column: int | None = None
    for index, header in enumerate(header_values):
        if normalize_header(header) in FINANCIAL_INSTITUTION_IN_COLUMNS:
            financial_institution_in_column = index
            continue
        mapped = map_header(header)
        if mapped:
            mapped_headers[index] = mapped

    mapped_fields = set(mapped_headers.values())
    missing = sorted(REQUIRED_ACCOUNT_FIELDS - mapped_fields)
    if missing:
        raise ValueError(
            "Workbook is missing required columns: " + ", ".join(missing)
        )
    if not {"account_status", "dormant_account"} & mapped_fields:
        raise ValueError(
            "Workbook is missing a required dormant status column. "
            "Use 'Account Status', 'Dormant', or 'IsDormant'."
        )

    records: list[AccountRecord] = []
    financial_institution_ins: set[str] = set()
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if (
            financial_institution_in_column is not None
            and financial_institution_in_column < len(row)
        ):
            raw_financial_institution_in = _text(
                row[financial_institution_in_column]
            )
            if raw_financial_institution_in:
                financial_institution_ins.add(
                    normalize_financial_institution_in(
                        raw_financial_institution_in,
                        source=(
                            "Uploaded workbook financial institution IN"
                        ),
                    )
                )
        values = {
            field: row[index] if index < len(row) else None
            for index, field in mapped_headers.items()
        }
        if is_summary_row(values):
            continue
        if not any(value not in (None, "") for value in values.values()):
            continue
        status_errors: list[str] = []
        account_status = False
        if "account_status" in values:
            account_status, account_status_error = normalize_account_status(
                values.get("account_status")
            )
            if account_status_error:
                status_errors.append(account_status_error)

        explicit_statuses: dict[str, bool] = {}
        for field, label in (
            ("dormant_account", "dormant"),
            ("closed_account", "closed"),
            ("undocumented_account", "undocumented"),
        ):
            if field not in values:
                continue
            normalized, status_error = normalize_status_indicator(
                values.get(field), label
            )
            explicit_statuses[field] = normalized
            if status_error:
                status_errors.append(status_error)

        dormant_account = explicit_statuses.get(
            "dormant_account", account_status
        )
        if (
            "account_status" in values
            and "dormant_account" in explicit_statuses
            and dormant_account != account_status
        ):
            status_errors.append(
                "Account Status and the dedicated dormant column conflict."
            )

        records.append(
            AccountRecord(
                row_number=row_number,
                account_number=_text(values["account_number"]),
                first_name=_text(values["first_name"]),
                surname=_text(values["surname"]),
                date_of_birth=normalize_date(
                    values["date_of_birth"], workbook.epoch
                ),
                address=_text(values["address"]),
                country=_text(values["country"]).upper(),
                tin=_text(values["tin"]),
                account_status=account_status,
                dormant_account=dormant_account,
                closed_account=explicit_statuses.get(
                    "closed_account", False
                ),
                undocumented_account=explicit_statuses.get(
                    "undocumented_account", False
                ),
                status_error=" ".join(status_errors),
                payment=_text(values["payment"]),
                account_balance=_text(values["account_balance"]),
            )
        )
    if len(financial_institution_ins) > 1:
        raise ValueError(
            "Uploaded workbook contains multiple financial institution IN "
            "values. Use one consistent value for the complete report."
        )
    return WorkbookParseResult(
        records=records,
        status_mapping=status_mapping_info(mapped_headers, header_values),
        financial_institution_in=next(
            iter(financial_institution_ins), None
        ),
    )


def parse_excel_bytes(content: bytes) -> list[AccountRecord]:
    return parse_excel_with_metadata(content).records
